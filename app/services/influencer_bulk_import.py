"""Bulk-import influencers from an uploaded spreadsheet (.xlsx) -- one row
per creator, with an optional Instagram and/or YouTube handle each, both
grouped under the same Creator via InfluencerRepo.create's existing
get-or-create-by-name behavior. See app/api/v1/admin.py's
POST /influencers/bulk and GET /influencers/bulk/template routes.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import CategoryNotFoundError, DuplicateInfluencerError
from app.repositories.category_repo import CategoryRepo
from app.repositories.influencer_repo import InfluencerRepo
from app.schemas.bulk_import import BulkImportResult, BulkImportRowResult
from app.schemas.influencer import InfluencerCreate

_VALID_ACCOUNT_TYPES = {"individual", "business"}


def _normalize_column_name(raw: str) -> str:
    return raw.strip().lower().replace(" ", "_")


@dataclass
class _ParsedRow:
    row_number: int
    creator_name: str
    category_name: str
    account_type: str
    instagram_handle: Optional[str]
    youtube_handle: Optional[str]
    # Set when the row fails validation before any DB work -- missing
    # required field, bad type value, or no handle at all.
    error: Optional[str] = None


def parse_bulk_import_rows(raw_rows: list[dict[str, Any]]) -> list[_ParsedRow]:
    """Pure function (no I/O, no DB access) so this is unit-testable
    without a real spreadsheet -- `raw_rows` is already column-name ->
    cell-value dicts (see read_bulk_import_workbook for the
    openpyxl-specific extraction). row_number is 1-based and counts data
    rows only (excludes the header row).

    Required columns: creator_name, category, type (individual/business).
    Optional: instagram_handle, youtube_handle -- at least one of the two
    must be present, since a row with neither has nothing to create."""
    parsed: list[_ParsedRow] = []
    for i, raw in enumerate(raw_rows, start=1):
        row = {
            _normalize_column_name(str(k)): (str(v).strip() if v is not None else "")
            for k, v in raw.items()
        }

        creator_name = row.get("creator_name", "")
        category_name = row.get("category", "")
        account_type_raw = row.get("type", "")
        # Leading "@" is optional/common on both platforms in casual use --
        # stripped here the same way AddInfluencerForm.jsx strips it
        # client-side for the single-add form, since InfluencerRepo.
        # normalize_handle only handles YouTube's URL/prefix forms, not a
        # bare "@" on either platform.
        instagram_handle = (row.get("instagram_handle") or "").lstrip("@") or None
        youtube_handle = row.get("youtube_handle") or None

        account_type_normalized = account_type_raw.lower()
        error: Optional[str] = None
        if not creator_name:
            error = "creator_name is required"
        elif not category_name:
            error = "category is required"
        elif not account_type_raw:
            error = "type is required (individual or business)"
        elif account_type_normalized not in _VALID_ACCOUNT_TYPES:
            error = f"type must be 'individual' or 'business', got '{account_type_raw}'"
        elif not instagram_handle and not youtube_handle:
            error = "at least one of instagram_handle or youtube_handle is required"

        parsed.append(
            _ParsedRow(
                row_number=i,
                creator_name=creator_name,
                category_name=category_name,
                account_type=account_type_normalized if account_type_normalized in _VALID_ACCOUNT_TYPES else "individual",
                instagram_handle=instagram_handle,
                youtube_handle=youtube_handle,
                error=error,
            )
        )
    return parsed


def read_bulk_import_workbook(file_bytes: bytes) -> list[dict[str, Any]]:
    """Reads the first worksheet of an uploaded .xlsx, treating row 1 as
    column headers -- returns one dict per non-blank data row, keyed by
    raw header text (parse_bulk_import_rows normalizes column names
    itself, so header capitalization/spacing doesn't matter here)."""
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []
    header_names = [str(h) if h is not None else "" for h in headers]

    result: list[dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue  # skip fully blank rows
        result.append(dict(zip(header_names, values)))
    return result


def build_bulk_import_template() -> bytes:
    """One-row example workbook matching the columns parse_bulk_import_rows
    expects -- served by GET /influencers/bulk/template."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Influencers"
    worksheet.append(["creator_name", "category", "type", "instagram_handle", "youtube_handle"])
    worksheet.append(["MrBeast", "Entertainment", "individual", "mrbeast", "@MrBeast"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def run_bulk_import(session: AsyncSession, raw_rows: list[dict[str, Any]]) -> BulkImportResult:
    """Creates an Influencer per handle present in each valid row (both
    grouped under one Creator when a row has both handles), via the same
    InfluencerRepo.create() the single-add form uses -- so bulk-imported
    rows behave identically to one entered by hand (same handle
    normalization, same get-or-create-by-name Creator linking, same
    duplicate-handle handling)."""
    parsed_rows = parse_bulk_import_rows(raw_rows)
    category_repo = CategoryRepo(session)
    influencer_repo = InfluencerRepo(session)

    results: list[BulkImportRowResult] = []
    created_count = 0
    error_count = 0

    for row in parsed_rows:
        if row.error:
            results.append(
                BulkImportRowResult(
                    row=row.row_number, creator_name=row.creator_name, status="error", message=row.error
                )
            )
            error_count += 1
            continue

        try:
            category = await category_repo.get_by_name(row.category_name)
        except CategoryNotFoundError:
            results.append(
                BulkImportRowResult(
                    row=row.row_number,
                    creator_name=row.creator_name,
                    status="error",
                    message=f"Category '{row.category_name}' not found",
                )
            )
            error_count += 1
            continue

        instagram_status: Optional[str] = None
        youtube_status: Optional[str] = None
        failure_messages: list[str] = []

        if row.instagram_handle:
            try:
                await influencer_repo.create(
                    InfluencerCreate(
                        handle=row.instagram_handle,
                        category_id=category.id,
                        platform="instagram",
                        creator_name=row.creator_name,
                        account_type=row.account_type,
                    )
                )
                instagram_status = "created"
            except DuplicateInfluencerError:
                instagram_status = "failed"
                failure_messages.append(f"Instagram @{row.instagram_handle} already registered")

        if row.youtube_handle:
            try:
                await influencer_repo.create(
                    InfluencerCreate(
                        handle=row.youtube_handle,
                        category_id=category.id,
                        platform="youtube",
                        creator_name=row.creator_name,
                        account_type=row.account_type,
                    )
                )
                youtube_status = "created"
            except DuplicateInfluencerError:
                youtube_status = "failed"
                failure_messages.append(f"YouTube {row.youtube_handle} already registered")

        attempted = [s for s in (instagram_status, youtube_status) if s is not None]
        succeeded = [s for s in attempted if s == "created"]

        if succeeded and len(succeeded) == len(attempted):
            status, message = "created", "Created successfully"
            created_count += 1
        elif succeeded:
            status, message = "partial", "; ".join(failure_messages)
            created_count += 1
        else:
            status, message = "error", "; ".join(failure_messages) or "Failed to create"
            error_count += 1

        results.append(
            BulkImportRowResult(
                row=row.row_number,
                creator_name=row.creator_name,
                status=status,
                message=message,
                instagram_handle=row.instagram_handle,
                instagram_status=instagram_status,
                youtube_handle=row.youtube_handle,
                youtube_status=youtube_status,
            )
        )

    return BulkImportResult(
        total_rows=len(parsed_rows),
        created_count=created_count,
        error_count=error_count,
        rows=results,
    )
